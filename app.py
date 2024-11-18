from flask import Flask, jsonify, render_template
import openpyxl
import random

app = Flask(__name__)

# Load data from the Excel file
file_path = 'data.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Convert Excel data into a list of dictionaries
flashcards = [
    {"question": sheet.cell(row=row, column=1).value,
      "answer": sheet.cell(row=row, column=2).value,
      "example": sheet.cell(row=row, column=3).value
      }
    for row in range(2, sheet.max_row + 1)
]

# Store the current flashcard index
current_index = random.randint(0, len(flashcards) - 1)

# Route to serve the index.html page
@app.route('/')
def index():
    return render_template('index.html')

# Endpoint to get a random question
@app.route('/get-question', methods=['GET'])
def get_question():
    global current_index
    current_index = random.randint(0, len(flashcards) - 1)
    return jsonify({"question": flashcards[current_index]["question"]})

# Endpoint to get the answer to the current question
@app.route('/get-answer', methods=['GET'])
def get_answer():
    return jsonify({"answer": flashcards[current_index]["answer"],"example": flashcards[current_index]["example"]})

if __name__ == '__main__':
    app.run(debug=True)
